'''파일 '신용거래내역조회_YYMMDD'에서 두번째 sheet의 붉은 색 row 정보 얻기
이전 마감 날짜의 오페라 카드 매출에 등록된 카드 신용거래내역중에, 당일 거래 내역이 있다면 붉은색으로 셋팅되어 있음.
따라서 이 색이 셋팅된 row 들을 획득하여, 그 중 '거래고유번호'를 반환한다
'''

from csv import excel
import openpyxl
import datetime
import os


def read_red_color() -> list:
    """신용거래내역에서 font가 빨간 라인의 '거래고유번호' 리스트를 반환한다
    2일 전의 신용카드 거래내역을 읽고, 그 중 두번쨰 sheet(거래내역들)을 활성 sheet로 셋팅.
    itter_rows를 이용한 for문을 통해 font의 color가 빨간색인것만('FFFF0000') 고르고, 그 줄의 첫번쨰 컬럼('거래고유번호')를 추출하여
    리스트 타입으로 반환한다
    
    Returns:
        list: {'거래고유번호' ...}
    """    
    # 1. 2일 전 마감된 신용카드 거래 내역 파일을 읽는다
    source_date = (datetime.datetime.now() - datetime.timedelta(2)).strftime("%y%m%d")      # 저장된 '신용거래내역_YYMMDD'용 날짜 포맷, 2일 전 날짜
    source_dir = 'C:/FA/creditcard/EDI_Confirm/'
    excel_filename = source_dir + f'신용거래내역조회_{source_date}.xlsx'                    # 2일 전 신용거래내역조회 파일
    
    try:
        if os.path.exists(excel_filename):
            edi_excel = openpyxl.load_workbook(excel_filename, data_only=False) # 수식파일 포함하여 엑셀파일 읽어 들임
            ws = edi_excel[edi_excel.sheetnames[1]]                             # 두번쨰 shhet 선택
        else:
            print('휴일인듯, 신용거래내역 없음')
            return (list())
    except Exception as e:
        with open('./error.log', 'a') as file:
            file.write(
                f'[font_red_on_excel.py - Reading Data] <{datetime.datetime.now()}> openpyxl file-reading error ({excel_filename}) ===> {e}\n'
            )
        raise(e)

    red_rows = []                           # 거래고유번호 리스트 
    for row in ws.iter_rows():              # row looping
        if row[0].font.color.rgb == 'FFFF0000':             # 해당 라인의 첫번쨰 셀의 font color가 red일 경우만.
            red_rows.append(row[0].value)                   # 해당 '거래고유번호'를 리스트에 추가
    
    return red_rows                         # 리스트 반환

if __name__ == '__main__':
    print(read_red_color())