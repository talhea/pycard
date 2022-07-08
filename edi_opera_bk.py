import pandas as pnds
import openpyxl
import datetime
import pickle

# 1. EDI 템플릿 엑셀 파일을 읽어서 DataFrame 생성
excel_filename = './EDI-xx월.xlsx'                                  # EDI 엑셀파일
edi_excel = openpyxl.load_workbook(excel_filename, data_only=False) # 수식파일 포함하여 엑셀파일 읽어 들임
ws = edi_excel.active                                               # 활성화된 sheet

yesterday = datetime.datetime.now() - datetime.timedelta(1) # 어제 날짜 추출
target_date = yesterday.strftime('%Y%m%d')                  # 어제 날짜 포맷

# 2. 오페라 어제 마감된 trial report(dt_trial_balance_날짜) dataframe 읽기
target_date = (datetime.datetime.now() - datetime.timedelta(1)).strftime("%Y%m%d")      # 어제 날짜 추출
dt_base_dir = './dtdata/' + target_date
data_dir = dt_base_dir + '/dt_opera_trial_' + target_date
with open(data_dir, "rb") as file:
    opera_df = pickle.load(file)

# 3. 오페라 dataframe의 trx_code와 EDI 엑셀 좌표 매칭 정보
map_code_edi = {
    # TRX_CODE : 엑셀좌표(A1, B13..)
    '9231': 'B7',
    '9232': 'B9',
    '9233': 'B8',
    '9234': 'B15',
    '9235': 'B6',
    '9236': 'B3',
    '9237': 'B12',
    '9238': 'B14',
    '9239': 'B10',
    '9240': 'B16',
    '9241': 'B17',
    '9242': 'B11',
    '9243': 'B5',
    '9244': 'B4',
    '9245': 'B15'
}

# 4. trial balance 내용을 순서대로 돌면서 엑셀에 입력
for code in opera_df.index:     # trx-code를 이용해서 ws와 opera_df에 접근하는 looping
    ws[map_code_edi[code]] = opera_df.loc[code]['TB_AMOUNT'] * -1       # 오페라 dataframe으로부터 tb_amount 정보를 가져와, 
                                                                        # 맵핑 딕셔너리에서 추출한 좌표를 이용해서 ws(EDI엑셀)에 넣는다
opera_date = yesterday.strftime('%b.%d.%Y') # 오페라 포맷용 어제 날짜
ws['B1'] = opera_date                       # EDI파일 날짜 셋팅
ws.title = str(yesterday.day)               # sheet 이름 변경

# 5. 파일 저장
excel_filename = dt_base_dir + '/EDI-' + target_date + '.xlsx'              # 파일 저장 시에 필요한 당 월의 엑셀파일 이름
edi_excel.save(excel_filename)