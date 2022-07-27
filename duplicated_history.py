'''이전(인수 날짜의 하루 전)의 거래내역에서 인수로 들어온 날(당일)로 된 '거래고유번호' 리스트를 반환한다
'''

from csv import excel
import openpyxl
import datetime
import os


def get_dup_serials(work_date) -> list:
    """이전(인수 날짜의 하루 전)의 거래내역에서 인수로 들어온 날(당일)로 된 '거래고유번호' 리스트를 반환한다
    '거래고유번호'는 당 일(day)을 기준으로 번호가 시작되므로, 이를 이용해서 중복된 당 일의 거래 내역(거래고유번호)을 추출해 반환한다
    
    Args:
        work_date (datetime): 어제 날짜
    
    Returns:
        list: {'거래고유번호' ...}
    """    
    # 1. 2일 전 마감된 신용카드 거래 내역 파일을 읽는다
    one_day_ago = (work_date - datetime.timedelta(1)).strftime('%y%m%d')       # '신용거래내역_YYMMDD'용 날짜 포맷, 하루 전(실제로는 2일) 날짜
    source_dir = 'C:/FA/creditcard/EDI_Confirm/'
    excel_filename = f'신용거래내역조회_{one_day_ago}.xlsx'
    
    #   '신용거래내역조회' 작업 파일 로딩
    try:
        if os.path.exists(source_dir + excel_filename):                         # 하루(실제로는 2일)전 신용거래내역조회 파일이 있을 경우
            edi_excel = openpyxl.load_workbook(source_dir + excel_filename, data_only=False)    # 수식파일 포함하여 엑셀파일 읽어 들임
            ws = edi_excel[edi_excel.sheetnames[1]]                                             # 두번쨰 sheet 선택
        else:
            print('duplicated_history.py : 휴일인듯, 신용거래내역 작업 파일 없음')
            return (list())                                                     # 빈 리스트 반환
    except Exception as e:
        with open('./error.log', 'a') as file:
            file.write(
                f'[tomorrow_history.py - Reading Data] <{datetime.datetime.now()}> openpyxl file-reading error ({excel_filename}) ===> {e}\n'
            )
        raise(e)

    # 2. 중복처리 되는 당일 처리 날짜의 거래고유번호를 찾아서 리스트 버퍼에 저장한다
    serial_rows = []                                # 거래고유번호 버퍼
    target_day = work_date.strftime('%d')           # 당일(어제)) 날짜의 일(day): 'DD' (두자리 문자열)
    
    #   엑셀의 'A'컬럼에서 각 값의 시작 글자를 확인해서 거래고유번호 추출
    for serial_num_string in ws['A']:               # 'A'컬럼('거래고유번호')  looping
        if serial_num_string.value != None:         # None 값으로 인해서 error 발생하는 경우 있음
            if serial_num_string.value.startswith(target_day) == True:          # 작업일 하루 전(2일 전) 거래내역에서의 '거래고유번호'의 시작이 작업일(어제) 날짜의 일(day)로 시작하는 경우
                serial_rows.append(serial_num_string.value)                     # 해당 거래는 현(당일) 작업 날짜와 중복되는 내용이므로, 이 떄의 '거래고유번호'를 리스트에 저장
    
    # 3. 중복 거래고유번호 리스트 반환
    return serial_rows                              # 리스트 반환


if __name__ == '__main__':
    print(get_dup_serials(datetime.datetime.now() - datetime.timedelta(1)))