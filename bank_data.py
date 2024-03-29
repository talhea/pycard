"""
다운로드 받은 기업은행, 농협(민국은행)의 입금 내역을 dataframe으로 저장한다
"""  

import pandas as pnds
from bs4 import BeautifulSoup
import pickle, os
import datetime
import re

def to_bank_df(work_date):
    """다운로드 받은 기업은행, 농협(민국은행)의 입금 내역을 dataframe으로 저장한다
    첫째, 기업은행 입금 내역을 readlines함수로 한 줄씩 읽어, 각 줄을 '|' 구분자로 다시 분리한 후 dataframe으로 만들고 전처리 과정을 거친다.
    둘째, 농협(민국은행) 입금 내역을 BeautifulSoup 객체로 읽어 Dataframe으로 전환해서 전 처리를 한다.
    세째, 두 데이터를 합쳐 저장한다
    
    Args:
        work_date (datetime): 오늘 날짜
    """    
    # 1. 기업은행에서 '텍스트형식저장'한 내역을 읽고 dataframe을 만든다
    target_date = (work_date - datetime.timedelta(1)).strftime("%Y%m%d")            # 어제 날짜 포맷
    receipts_date = work_date.strftime("%Y%m%d")                                    # 입금 오늘 날짜
    
    downdata_dir = f'./data/{target_date}/downdata/'                                # 읽어들일 down디렉토리 './data/YYYYMMDD/downdata/'
    ibk_filename = '거래내역조회_입출식 예금' + receipts_date  + '.txt'               # IBK은행 입금내역 파일 '거래내역조회_입출식 예금YYYYMMDD.txt'
    
    # 1-1. .txt 파일을 한 줄씩 읽어서 리스트에 넣는다.
    if os.path.exists(downdata_dir + ibk_filename) == True:                         # 기업은행 입금 내역이 있을 경우
        #   입금내역 txt파일을 라인 단위로 읽어들인다
        try:
            with open(downdata_dir + ibk_filename, 'r', encoding='euc-kr') as file:
                lines = file.readlines()
        except Exception as e:
            with open('./error.log', 'a') as file:
                file.write(
                    f'[bank_data.py - Reading Data] <{datetime.datetime.now()}> file-reading error ({ibk_filename}) ===> {e}\n'
                )
            raise(e)
        
        # 1-2. 리스트의 각 라인을 '|' 기준으로 분리해서 dataframe을 만든다
        lines = list(map(lambda line: line.strip().split('|'), lines))      # '\n' 제거하고 '|' 단위로 분해
        
        ibk_df = pnds.DataFrame(lines)                                      # DataFrame 생성
        ibk_df.rename(columns=ibk_df.iloc[0], inplace=True)                 # 첫번쨰 행을 컬럼이름으로 지정
        ibk_df.drop(ibk_df.index[0], inplace=True)                          # 위 라인 실행결과로 중복되는 첫번째 행 삭제

        # 2. 데이터 전처리
        # 2-1. 필요 컬럼 추출하고 컬럼명 변경
        ibk_df = ibk_df[['거래일시', '입금', '거래내용', '상대계좌예금주명']]   # 필요한 컬럼 추출
        ibk_df.columns = ['date', 'receipts', 'details', 'holder']

        # 2-2. 금액 컬럼의 값에서 천단위 구분자 ','를 제외해서 int형으로 타입 변경
        ibk_df['receipts'] = ibk_df['receipts'].str.replace(',', '').astype('int64', errors='ignore')

        # 2-3. holder 데이터를 카드사 이름으로 셋팅 : HD, LT, SS, SH, KEB, BC, KB (NH는 따로 셋팅)
        ibk_df = ibk_df.replace({'holder': '현대카드（주）'}, 'HD')
        ibk_df = ibk_df.replace({'holder': '롯데카드（주）'}, 'LT')
        ibk_df = ibk_df.replace({'holder': '삼성카드（주）'}, 'SS')
        ibk_df = ibk_df.replace({'holder': '신한카드（주）'}, 'SH')
        ibk_df = ibk_df.replace({'holder': '하나카드　주식회사'}, 'KEB')
        ibk_df = ibk_df.replace({'holder': '비씨카드（주）'}, 'BC')                 # 해외카드 승인된 BC가 입금될때

        # 2-4. details 컬럼에 'BC' 문자열 포함한 행들 추출해서 'holder' 컬럼 값 변경    # 국내 승인된 BC가 입금될떄
        bc_expr = "details.str.endswith('BC')"                  # BC로 끝나는 문자열
        bc_lst = ibk_df.query(bc_expr).index.tolist()           # 조건 부합하는 열 값을 가진 행을 추출
        ibk_df.loc[bc_lst, 'holder'] = 'BC'                     # bc_lst 리스트에 포함된 모든 행의 'holder'컬럼 값 'KB'로 변경

        # 2-5. details 컬럼에 'KB' 문자열 포함한 행들 추출해서 'holder' 컬럼 값 변경
        kb_expr = "details.str.startswith('KB')"                # KB로 시작하는 문자열
        kb_lst = ibk_df.query(kb_expr).index.tolist()           # 조건 부합하는 열 값을 가진 행을 추출
        ibk_df.loc[kb_lst, 'holder'] = 'KB'                     # kb_lst 리스트에 포함된 모든 행의 'holder'컬럼 값 'KB'로 변경

        # 3. 카드사 입금 거래 내역만 추출
        ibk_df = ibk_df[ibk_df['holder'].isin(['HD', 'LT', 'SS', 'SH', 'KEB', 'BC', 'KB'])]
    else:
        # 기업은행 입금 내역 없는 경우
        ibk_df = pnds.DataFrame(columns=range(5))               # 빈 Dataframe 생성

    #------------------------------------------------------------------------------------------------------------------
    # 기업은행 끝, 농협 시작 : 'NH' 셋팅
    #------------------------------------------------------------------------------------------------------------------

    # 4. 농협의 xml 파일은 실제 내용은 html, 따라서 BeautifulSoup를 이용해서 읽어들임
    nh_filename = receipts_date + '.xls'            # 읽을 파일 이름 'YYYYMMDD.xls'
    
    #   BeautifulSoup 객체로 읽어들임
    try:
        with open(downdata_dir + nh_filename, 'rt', encoding='UTF-8') as page:
            soup = BeautifulSoup(page, 'html.parser')
    except Exception as e:
        with open('./error.log', 'a') as file:
            file.write(
                f'[bank_data.py - Reading Data] <{datetime.datetime.now()}> file-reading error ({nh_filename}) ===> {e}\n'
            )
        # 기능에러가 아닌 빈 거래내역 파일 자체가 없는 경우에도 문제이므로 프로그램 종료
        raise(e)

    # 5. dataframe 생성
    # 5-1. 내용 추출
    content_soup = soup.select('td.se-td')      # 입출금내역을 추출: 한 라인의 각 컬럼을 리스트 형식으로 저장
    row_lst = []                                # DataFrame 생성에 사용할 line buffer, 컬렴명이나 인텍스 없는 내용만 저장
    
    for content in content_soup:                # 각 컬럼을 looping
        row_lst.append(re.sub('([0-9,]+) 원', '\\1', content.get_text().strip()))   # text 주변의 공백을 제거한 내용만 추출

    if len(content_soup) >= 5:                  # 입금 내역이 있을때만: 입금 내역이 있다면 최소한 5 이상의 길이(컬럼 수가 5개)가 나온다
        # 5-2. content 리스트를 dataframe으로 변환 ['날짜', '출금', '입금', '잔액', '적요']
        nh_df = pnds.DataFrame([row_lst], columns=['date', 'drawing', 'receipts', 'balance', 'holder'])

        # 6. data 전처리
        # 6-1. 날짜 포맷 : '%Y-%m-%d %H:%M:%S' => 두 dataframe 통합 후에 정렬 및 날짜포맷 적용을 위함
        nh_df['date'] = nh_df['date'].astype('datetime64[ns]', errors='ignore')     # date 포맷 변경을 위해서 datetime 형으로 변환
        nh_df['date'] = nh_df['date'].dt.strftime('%Y-%m-%d %H:%M:%S')              # 연산을 통해 포맷 변경 => 반환 타입은 일반 객체로 변경됨

        # 6-2. 천단위 ','를 없애고 금액 타입을 'int64'로 변경
        nh_df['receipts'] = nh_df['receipts'].str.replace(',', '').astype('int64', errors='ignore')

        # 6-3. 'details' 컬럼 추가하고, 필요한 컬럼만 추출
        nh_df['details'] = 'NH11381135'
        nh_df = nh_df[['date', 'receipts', 'details', 'holder']]

        # 6-4. NH카드에서 입금된 라인의 holder 값을 'NH'로 바꿈
        str_expr = "holder.str.startswith('NH11381135')"        # NH로 시작하는 문자열
        nh_lst = nh_df.query(str_expr).index.tolist()           # 조건 부합하는 열 값을 가진 행을 추출
        nh_df.loc[nh_lst, 'holder'] = 'NH'                      # nh_lst 리스트에 있는 행들의 'holder'컬럼 값을 'NH'로 변경

        # 7. NH카드 입금 라인만 추출
        nh_df = nh_df[nh_df['holder'].isin(['NH'])]
    else:
        # NH 입금내역이 없는 경우
        nh_df = pnds.DataFrame(columns=range(5))                # 빈 Dataframe 생성

    #------------------------------------------------------------------------------------------------------------------
    # 농협 끝, 기업은행과 눙협 합치기
    #------------------------------------------------------------------------------------------------------------------

    # 8. 기업은행 데이타프레임(ibk_df)와 농협 데이타프레임(nh_df)를 합친다
    if (ibk_df.empty == False) & (nh_df.empty == True):         # 기업은행만 거래내역 있을 경우
        bank_df = ibk_df
    elif (ibk_df.empty == False) & (nh_df.empty == False):      # 기업/농협 모두 거래내역 존재하면 Dataframe 합침
        bank_df = pnds.concat([ibk_df, nh_df], ignore_index=True)
    elif (nh_df.empty == False) & (ibk_df.empty == True):       # 농협만 거래내역 있을 경우
        bank_df = nh_df
    else:
        # 기업/농협 모두 거래내역 없으면 프로그램 종료
        print('기업은행, 농협 거래내역 모두 없음')
        exit()
    
    # 9. data 전처리
    # 9-1. 거래일시 기준 Sorting
    bank_df = bank_df.sort_values(by=['date'])

    # 9-2. 날짜 포맷 : '%Y-%m-%d'
    bank_df['date'] = bank_df['date'].map(lambda str_data: str_data.split()[0], na_action='ignore') # 날짜 포맷 : '%Y-%m-%d'
    
    # 10. dfdata 디렉토리에 저장
    dfdata_dir = f'./data/{target_date}/dfdata/'        # 데이타를 저장할 df디렉토리 './data/YYYYMMDD/dfdata/'
    
    # 10-1. 목적지 df디렉토리 확인하고 없으면 생성
    try:
        if os.path.exists(dfdata_dir) == False:         # 폴더가 없으면 생성
            os.makedirs(dfdata_dir)
    except Exception as e:
        with open('./error.log', 'a') as file:          # error 로그 파일에 추가
            file.write(
                f'[bank_data.py - Making dfdata] <{datetime.datetime.now()}> mkdir error ({dfdata_dir}) ===> {e}\n'
            )
        raise(e)
    
    # 10-2. dataframe 저장
    try:
        df_filename = 'df_bank_' + receipts_date        # 저장할 dataframe 파일 이름 'df_bank_YYYYMMDD'

        with open(dfdata_dir + df_filename, "wb") as file:
            pickle.dump(bank_df, file)
    except Exception as e:
        with open('./error.log', 'a') as file:
            file.write(
                f'[bank_data.py - Making Dataframe] <{datetime.datetime.now()}> pickle.dump error ({df_filename}) ===> {e}\n'
            )
        raise(e)
    
    # 10-3. KICC 입금현황 엑셀 파일에 Bank Dataframe을 저장
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment
    import openpyxl

    #   KICC 입금현황 엑셀 파일 로딩
    kicc_receipts_filename = './downdata/20220630/입금현황 · 일별 (2)_kicc.xlsx'
    
    bank_wb = openpyxl.load_workbook(downdata_dir + kicc_receipts_filename)
    bank_sheet = bank_wb.active
    bank_sheet.title = f'{receipts_date}'
    row_count = bank_sheet.max_row                      # 추가할 Dataframe의 첫 라인을 강조('Pandas' style)하기 위한 행 참조
    
    #   필요없는 컬럼 삭제
    bank_sheet.delete_cols(2, 3)                        # 삭제 컬럼: 사업자등록번호, 사업자상호, 가맹점번호
    
    #   Dataframe을 추가하기 전에 엑셀 파일에 맞게 컬럼 조정
    bank_df = bank_df[['holder', 'date', 'details', 'receipts']]                # 컬럼 순서 재조정
    for idx in range(6):                                                        # 컬럼 갯수 조정
        bank_df.insert(2, '--', '', allow_duplicates=True)
    
    #   openpyxl에 Datframe을 주입
    for row in dataframe_to_rows(bank_df, index=False, header=True):            # index 제외
        bank_sheet.append(row)

    #   각 컬럼의 width 설정
    for col in range(97, 97 + bank_sheet.max_column):                           # 영문 'A'의 십진수 97...
        bank_sheet.column_dimensions[chr(col)].width = 13.6
    
    #   각 행의 height 설정
    for row in range(bank_sheet.max_row):
        bank_sheet.row_dimensions[row].height = 18

    #   column 이름을 강조하는 'Pandas' style
    for cell in bank_sheet[1] + bank_sheet[row_count + 1]:                      # KICC 입금현황 첫 줄과 추가한 Dataframe 첫 줄
        cell.style = 'Pandas'

    #   추가한 Dataframe을 위하여 해당 컬럼 재정렬
    for cell in bank_sheet['A'] + bank_sheet['B'] + bank_sheet['I']:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    #   'J' Dataframe 컬럼(입금액, receipts)을 위한 숫자 포맷 셋팅
    for cell in bank_sheet['J']:
        cell.number_format = '#,### '
    
    #   엑셀 저장
    try:
        target_dir = f'./data/{target_date}/'           # 저장 위치 './data/YYYYMMDD'
        xl_filename = df_filename + '.xlsx'             # 저장파일 'df_bank_YYYYMMDD.xlsx'
        
        bank_wb.save(target_dir + xl_filename)
    except Exception as e:
        with open('./error.log', 'a') as file:
            file.write(
                f'[kicc_history.py - Making Excel] <{datetime.datetime.now()}> openpyxl save() error ({xl_filename}) ===> {e}\n'
            )
        raise(e)


if __name__ == '__main__':
    to_bank_df(datetime.datetime.now())