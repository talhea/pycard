'''파일 '신용거래내역조회_YYMMDD'에서 두번째 sheet의 붉은 색 row 정보 얻기
이전 마감 날짜의 오페라 카드 매출에 등록된 카드 신용거래내역중에, 당일 거래 내역이 있다면 붉은색으로 셋팅되어 있음.
따라서 이 색이 셋팅된 row 들을 획득하여, 그 중 '거래고유번호'를 반환한다

'''

from csv import excel
import openpyxl
import datetime
import os


def get_serial_number(work_date) -> list:
    """신용거래내역에서 font가 빨간 라인의 '거래고유번호' 리스트를 반환한다
    '거래고유번호'는 당 일(day)을 기준으로 번호가 시작된다
    '거래고유번호'는 당 일(day)을 기준으로 번호가 시작된다
    
    Args:
        work_date (datetime): 어제 날짜
    
    Returns:
        list: {'거래고유번호' ...}
    """    
    # 1. 2일 전 마감된 신용카드 거래 내역 파일을 읽는다
    source_date = (work_date - datetime.timedelta(1)).strftime('%y%m%d')        # '신용거래내역_YYMMDD'용 날짜 포맷, 2일 전 날짜
    source_dir = 'C:/FA/creditcard/EDI_Confirm/'
    excel_filename = f'신용거래내역조회_{source_date}.xlsx'
    
    try:
        if os.path.exists(source_dir + excel_filename):                     # 2일 전 신용거래내역조회 파일이 있을 경우
            edi_excel = openpyxl.load_workbook(source_dir + excel_filename, data_only=False)    # 수식파일 포함하여 엑셀파일 읽어 들임
            ws = edi_excel[edi_excel.sheetnames[1]]                                             # 두번쨰 sheet 선택
        else:
            print('read_red_coloe.py : 휴일인듯, 신용거래내역 없음')
            return (list())
    except Exception as e:
        with open('./error.log', 'a') as file:
            file.write(
                f'[font_red_on_excel.py - Reading Data] <{datetime.datetime.now()}> openpyxl file-reading error ({excel_filename}) ===> {e}\n'
            )
        raise(e)

    red_rows = []                               # 거래고유번호 리스트 
    tommorow = work_date.strftime('%d')         # 내일 날짜의 일(day)
    
    for serial_num_string in ws['A']:               # '거래고유번호' 컬럼 looping
        if serial_num_string.value.startswith(tommorow):    # '거래고유번호'의 시작이 내일날짜의 일(day)로 시작하는 경우
            red_rows.append(serial_num_string.value)        # 해당 거래는 내일 날짜 거래이므로, 이 떄의 '거래고유번호'를 리스트에 저장
    
    return red_rows                         # 리스트 반환

if __name__ == '__main__':
    print(read_red_color())